import pathlib
import re

import openpyxl


class ExcelDictionaryFile:
    """
    辞書検索出力用のエクセルファイル
    """
    # ヘッダーに使う文字（辞書のキーとしても使用）
    HEADER_WORDS = ['単語', '反意語', '類義語', '派生語', '発音（英）', '発音（米）']

    # ハイパーリンクを設定するヘッダー
    HYPERLINK_HEADER_WORDS = ['発音（英）', '発音（米）']

    def __init__(self, file_name):
        # Windowsファイルで使えない文字の置き換え
        file_name = re.sub(r'[\\/:*?"<>|]+', '_', file_name)
        self._path = pathlib.Path(file_name)

    def write(self, words):
        """書き込み
        Args:
            words (list[dict]): 単語の辞書のリスト
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        # ヘッダー書き込み
        for i, header_word in enumerate(ExcelDictionaryFile.HEADER_WORDS):
            ws.cell(row=1, column=i+1).value = header_word

        # 単語書き込み
        row = 2
        for word_dict in words:

            # ヘッダーの文字をキーに辞書から単語を取得
            for i, key in enumerate(ExcelDictionaryFile.HEADER_WORDS):
                ws.cell(row=row, column=i+1).value = word_dict[key]

                # ハイパーリンクを設定するヘッダーであれば設定
                if key in ExcelDictionaryFile.HYPERLINK_HEADER_WORDS:
                    ws.cell(row=row, column=i+1).hyperlink = word_dict[key]

            # 次の行に移動してループ
            row += 1

        # 保存
        wb.save(str(self._path))
